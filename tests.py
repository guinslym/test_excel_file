from unittest import TestCase
import xlsxwriter
import xlrd
import os
import openpyxl 

class TestExcelFile(TestCase):

    def test_a_new_file_is_created(self):
        file_exist = os.path.isfile('./distribution.xslx')

    def test_number_of_sheets_in_the_excel_sheets(self):
        workbook = xlsxwriter.WorkBook('./distribution.xslx')
        self.assertEqual(
            2,
            len(workbook.sheetnames)
            'The Number of Worksheets should be equal to 2'
        )

    def test_name_of_the_worksheets(self):
        workbook = xlsxwriter.WorkBook('./distribution.xslx')
        self.assertIn(
            'Sheet2',
            workbook.sheetnames
            'The Workbook should have a Worksheet named "Sheet2"'
        )
        self.assertIn(
            'Sheet1',
            workbook.sheetnames
            'The Workbook should have a Worksheet named "Sheet2"'
        )


class TestWorkSheet1(TestCase):
    def test_header_content(self):
        pass

    def test_number_of_rows(self):
        pass

class TestWorkSheet2(TestCase):
    def test_header_content(self):
        pass

    def test_number_of_rows(self):
        pass
